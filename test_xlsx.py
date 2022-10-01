import openpyxl

def init_exls_file(filename,sheetname1,sheetname2,sheetname3,sheetname4):
    excel_file = openpyxl.load_workbook(filename)
    sheet_object1 = excel_file[sheetname1]
    sheet_object2 = excel_file[sheetname2]
    sheet_object3 = excel_file[sheetname3]
    sheet_object4 = excel_file[sheetname4]
    return excel_file,sheet_object1,sheet_object2,sheet_object3,sheet_object4

def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows

def delete_blank_rows(*, sheet_object):
    y = 1
    while (y < sheet_object.max_row):
        flag = 0
        for x in range(1, sheet_object.max_column + 1):
            if sheet_object.cell(row=y, column=x).value == None:
                flag= 1
            else: flag = 0
        if(flag == 1):
            sheet_object.delete_rows(idx=y, amount=1)
            y-=1
        y+=1
    excel_file.save('client.xlsx')

def add_client(new_client,*, sheet_object,excel_file):
    sheet_object.append(new_client)
    excel_file.save('client.xlsx')

def unike_test(new_client,*, sheet_object):
    flag = 0
    for y in range(2,sheet_object.max_row+1):
        if sheet_object.cell(row=y, column=1).value == new_client[0]: flag  = 1
    return flag

def show_all_clients(*, sheet_object):
    for y in range(2,sheet_object.max_row+1):
        for x in range(1, sheet_object.max_column+1):
            print(sheet_object.cell(row=y, column=x).value, end= ' ')
        print()

def write_row_txt(*, sheet_object):
    ms = ''
    for x in range(1, sheet_object.max_column+1):
        ms = ms + str(sheet_object.cell(row=sheet_object.max_row, column=x).value) + ' | '
    ms += '\n'
    with open("client.txt", "at", encoding='utf-8') as file:
        file.write(ms)

def write_all_txt(*, sheet_object):
    msg = ''
    for y in range(2,sheet_object.max_row+1):
        for x in range(1, sheet_object.max_column+1):
            msg = msg + str(sheet_object.cell(row=y, column=x).value) + ' | '
        msg += '\n'
    with open("client.txt", "at", encoding='utf-8') as file:
        file.write(msg)

def form_req(*, sheet_object,excel_file,mich_sheet,manya_sheet,artem_sheet):
    mich_sheet.delete_rows(idx=1, amount=mich_sheet.max_row)
    manya_sheet.delete_rows(idx=1, amount=manya_sheet.max_row)
    artem_sheet.delete_rows(idx=1, amount=artem_sheet.max_row)
    goal = 1
    for y in range(2,sheet_object.max_row+1):
        client = []
        for x in range(1, sheet_object.max_column+1):
            client.append(sheet_object.cell(row=y, column=x).value)
        if (goal == 1) and (client[8] != 'оплачено'): mich_sheet.append(client)
        elif (goal == 2) and (client[8] != 'оплачено'): manya_sheet.append(client)
        elif (client[8] != 'оплачено'):
            artem_sheet.append(client)
            goal = 0
        goal += 1
    excel_file.save('client.xlsx')

def get_req(dropper_id,*,mich_sheet,manya_sheet,artem_sheet):
    text = ''
    if (dropper_id == 1):
        for y in range(1,mich_sheet.max_row+1):
            for x in range(1, mich_sheet.max_column+1):
                if (x in [1,4,5,6,9,11]): text = text + str(mich_sheet.cell(row=y, column=x).value) + ' | '
            text += '\n'
            text += '\n'
    if (dropper_id == 2):
        for y in range(1,manya_sheet.max_row+1):
            for x in range(1, manya_sheet.max_column+1):
                if (x in [1,4,5,6,9,11]): text = text + str(manya_sheet.cell(row=y, column=x).value) + ' | '
            text += '\n'
            text += '\n'
    if (dropper_id == 3):
        for y in range(1,artem_sheet.max_row+1):
            for x in range(1, artem_sheet.max_column+1):
                if (x in [1,4,5,6,9,11]): text = text + str(artem_sheet.cell(row=y, column=x).value) + ' | '
            text += '\n'
            text += '\n'
    return text

def get_rate(*,sheet_object):
    prom_name =[]
    prom = []
    msg = ''
    for y in range(2,sheet_object.max_row+1):
        i = sheet_object.cell(row=y, column=8).value
        prom.append(i)
        if i not in prom_name: prom_name.append(i)
    prom_name.sort(key=lambda n: prom.count(n),reverse=True)
    for i in prom_name: msg = msg + i + ' | ' + str(prom.count(i)) + "\n"
    return (msg)

#-------LESGO----------------------------

excel_file,client_sheet,mich_sheet,manya_sheet,artem_sheet = init_exls_file('client.xlsx','all','Михей','Маня','Артём')

#delete_blank_rows(sheet_object=client_sheet)

#print(f'Total Rows = {client_sheet.max_row} and Total Columns = {client_sheet.max_column}')

#show_all_clients(sheet_object=client_sheet)

#new_client = ['#7817435', 'билеты',	'оплачено', '+7(960)252-79-24', '19dn@outlook.com', 'Савин Гарри Федотович', '03.09.2022 10:43', 'Артём', 'бабло на карту']

#print(form_req(1,sheet_object=client_sheet))

#print(unike_test(new_client, sheet_object=client_sheet))
#add_client(new_client, sheet_object=client_sheet)
#write_all_txt(sheet_object=client_sheet)
#print(get_rate(sheet_object=client_sheet))